#conda create -n Vitale python=3.9
#conda activate Vitale
#cd D:\Woobsing\Vitale

import pandas as pd
import numpy as np
import pandas_profiling as pf

# #INFORME PROFILE######################################################################################
# #INFORME PROFILE######################################################################################
# #INFORME PROFILE######################################################################################
# df = pd.read_excel('BD.xlsx')
# # rpt = pf.ProfileReport(df)
# # rpt.to_file('profile_report.html')

# #INFORME SWEETVIZ######################################################################################
# #INFORME SWEETVIZ######################################################################################
# #INFORME SWEETVIZ######################################################################################
# import sweetviz as sv
# reporte = sv.analyze(df)

# reporte.show_html()


##################################################################################################################
##################################################################################################################
##################################################################################################################

from collections import Counter
from string import punctuation
import re
from openpyxl import load_workbook

df = pd.read_excel('BD.xlsx')

ENFERMEDAD_ACTUAL = df['ENFERMEDAD ACTUAL'].tolist() 
# ANTECEDENTES
# EXAMEN_FISICO
# IMPRESION_ANALISIS
# PLAN_DE_MANEJO

strips = str(ENFERMEDAD_ACTUAL).split(" ")
text = str(ENFERMEDAD_ACTUAL)
strips = re.findall(r"[\w]+|[.,!?;: ]", text)
print(strips)
# c = Counter((x.rstrip(punctuation).lower() for y in strips for x in y.split()))

# df1 = pd.DataFrame.from_dict(c, orient='index').reset_index()
# df1 = df1.rename(columns={'index':'Word', 0:'Frec.'})
# df1['Columna'] = 'ENFERMEDAD ACTUAL'
# print(df1)




# rows = df1.values.tolist()
# workbook = load_workbook(filename="output.xlsx")
# sheet = workbook["Sheet1"]
# for row in rows:
#     sheet.append(row)
# workbook.save(filename="output.xlsx")
# workbook.close

# ANTECEDENTES = df['ANTECEDENTES'].tolist() 
# text1 = str(ANTECEDENTES)
# strips1 = re.findall(r"[\w']+|[.,!?;:]", text1)
# c1 = Counter((x.rstrip(punctuation).lower() for y in strips1 for x in y.split()))
# df2 = pd.DataFrame.from_dict(c1, orient='index').reset_index()
# df2 = df2.rename(columns={'index':'Word', 0:'Frec.'})

# df2['Columna'] = 'ANTECEDENTES'
# rows = df2.values.tolist()
# workbook = load_workbook(filename="output.xlsx")
# sheet = workbook["Sheet1"]
# for row in rows:
#     sheet.append(row)
# workbook.save(filename="output.xlsx")
# workbook.close

# EXAMEN_FISICO = df['EXAMEN FISICO'].tolist() 
# text2 = str(EXAMEN_FISICO)
# strips2 = re.findall(r"[\w']+|[.,!?;:]", text2)
# c2 = Counter((x.rstrip(punctuation).lower() for y in strips2 for x in y.split()))
# df3 = pd.DataFrame.from_dict(c2, orient='index').reset_index()
# df3 = df3.rename(columns={'index':'Word', 0:'Frec.'})

# df3['Columna'] = 'EXAMEN FISICO'
# rows = df3.values.tolist()
# workbook = load_workbook(filename="output.xlsx")
# sheet = workbook["Sheet1"]
# for row in rows:
#     sheet.append(row)
# workbook.save(filename="output.xlsx")
# workbook.close

# IMPRESION_ANALISIS = df['IMPRESION (ANALISIS)'].tolist() 
# text3 = str(IMPRESION_ANALISIS)
# strips3 = re.findall(r"[\w']+|[.,!?;:]", text3)
# c3 = Counter((x.rstrip(punctuation).lower() for y in strips3 for x in y.split()))
# df4 = pd.DataFrame.from_dict(c3, orient='index').reset_index()
# df4= df4.rename(columns={'index':'Word', 0:'Frec.'})

# df4['Columna'] = 'IMPRESION (ANALISIS)'
# rows = df4.values.tolist()
# workbook = load_workbook(filename="output.xlsx")
# sheet = workbook["Sheet1"]
# for row in rows:
#     sheet.append(row)
# workbook.save(filename="output.xlsx")
# workbook.close


# PLAN_DE_MANEJO = df['PLAN DE MANEJO '].tolist() 
# text4 = str(PLAN_DE_MANEJO)
# strips4 = re.findall(r"[\w']+|[.,!?;:]", text4)
# c4 = Counter((x.rstrip(punctuation).lower() for y in strips4 for x in y.split()))
# df5 = pd.DataFrame.from_dict(c4, orient='index').reset_index()
# df5= df5.rename(columns={'index':'Word', 0:'Frec.'})

# df5['Columna'] = 'PLAN DE MANEJO '
# rows = df5.values.tolist()
# workbook = load_workbook(filename="output.xlsx")
# sheet = workbook["Sheet1"]
# for row in rows:
#     sheet.append(row)
# workbook.save(filename="output.xlsx")
# workbook.close






